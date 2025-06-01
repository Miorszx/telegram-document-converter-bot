"""
Document conversion functionality for the Telegram Document Converter Bot
"""

import os
import io
import logging
import asyncio
import uuid
import shutil
import subprocess
from datetime import datetime
from typing import List
from concurrent.futures import ThreadPoolExecutor

from PIL import Image, ImageEnhance, ImageFilter
import fitz  # PyMuPDF
from docx import Document as DocxDocument
import openpyxl
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Image as RLImage, PageBreak, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.lib import colors

from config.config import BotConfig


class DocumentConverter:
    """Enhanced document conversion functionality with proper Excel formatting and custom naming"""
    
    def __init__(self, config: BotConfig):
        self.config = config
        self.executor = ThreadPoolExecutor(max_workers=config.max_concurrent_conversions)
        self.logger = logging.getLogger(__name__)
    
    async def images_to_pdf(self, image_paths: List[str], output_path: str, 
                           quality: str = 'medium', add_metadata: bool = True) -> bool:
        """Convert multiple images to a single PDF with fixed scaling"""
        try:
            loop = asyncio.get_event_loop()
            return await loop.run_in_executor(
                self.executor, 
                self._images_to_pdf_sync, 
                image_paths, output_path, quality, add_metadata
            )
        except Exception as e:
            self.logger.error(f"Error converting images to PDF: {e}")
            return False
    
    def _images_to_pdf_sync(self, image_paths: List[str], output_path: str, 
                           quality: str, add_metadata: bool) -> bool:
        """Fixed synchronous image to PDF conversion"""
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.platypus import SimpleDocTemplate, Image as RLImage, PageBreak
            
            # Create PDF with standard margins
            doc = SimpleDocTemplate(
                output_path, 
                pagesize=A4,
                rightMargin=36,  # 0.5 inch
                leftMargin=36,
                topMargin=36,
                bottomMargin=36
            )
            
            elements = []
            
            # Get usable page dimensions
            page_width, page_height = A4
            margin = 36
            max_width = page_width - 2 * margin  # 523.27
            max_height = page_height - 2 * margin  # 806.89
            
            self.logger.info(f"PDF page dimensions: {page_width}x{page_height}")
            self.logger.info(f"Available space: {max_width}x{max_height}")
            
            for i, img_path in enumerate(image_paths):
                try:
                    # Verify image exists
                    if not os.path.exists(img_path):
                        self.logger.error(f"Image file not found: {img_path}")
                        continue
                    
                    # Clean up any enhanced image chains to avoid long filenames
                    if 'enhanced' in img_path and len(os.path.basename(img_path)) > 50:
                        # Create a clean copy
                        base_dir = os.path.dirname(img_path)
                        clean_name = f"clean_img_{i}_{uuid.uuid4().hex[:8]}.jpg"
                        clean_path = os.path.join(base_dir, clean_name)
                        
                        # Copy and optimize the image
                        with Image.open(img_path) as img:
                            if img.mode not in ['RGB', 'L']:
                                img = img.convert('RGB')
                            img.save(clean_path, 'JPEG', quality=85, optimize=True)
                        
                        img_path = clean_path
                    
                    # Get image dimensions
                    with Image.open(img_path) as img:
                        img_width, img_height = img.size
                        self.logger.info(f"Image {i+1}: {img_width}x{img_height}")
                        
                        # Calculate scale to fit within page bounds with safety margin
                        scale_x = (max_width * 0.95) / img_width  # 5% safety margin
                        scale_y = (max_height * 0.95) / img_height
                        scale = min(scale_x, scale_y, 1.0)  # Never upscale
                        
                        # Calculate final dimensions
                        final_width = img_width * scale
                        final_height = img_height * scale
                        
                        # Ensure we don't exceed bounds (double safety check)
                        if final_width > max_width:
                            final_width = max_width * 0.9
                            final_height = (final_width / img_width) * img_height
                        
                        if final_height > max_height:
                            final_height = max_height * 0.9
                            final_width = (final_height / img_height) * img_width
                        
                        self.logger.info(f"Final size: {final_width}x{final_height}")
                        
                        # Add image to PDF with fixed dimensions
                        rl_img = RLImage(img_path, width=final_width, height=final_height)
                        elements.append(rl_img)
                        
                        # Add page break except for last image
                        if i < len(image_paths) - 1:
                            elements.append(PageBreak())
                            
                except Exception as img_error:
                    self.logger.error(f"Error processing image {i+1}: {img_error}")
                    # Continue with other images
                    continue
            
            if elements:
                # Build PDF
                doc.build(elements)
                self.logger.info(f"PDF successfully created: {output_path}")
                
                # Add metadata if requested
                if add_metadata:
                    self._add_pdf_metadata(output_path)
                
                return True
            else:
                self.logger.error("No images could be processed")
                return False
            
        except Exception as e:
            self.logger.error(f"Error in _images_to_pdf_sync: {e}")
            return False
    
    def _add_pdf_metadata(self, pdf_path: str):
        """Add metadata to PDF"""
        try:
            doc = fitz.open(pdf_path)
            metadata = {
                "title": "Converted Images",
                "author": "TelegramBot",
                "creator": "Advanced Document Converter",
                "creationDate": datetime.now(),
                "modDate": datetime.now()
            }
            doc.set_metadata(metadata)
            doc.save(pdf_path, incremental=True)
            doc.close()
        except Exception as e:
            self.logger.warning(f"Could not add metadata: {e}")
    
    async def pdf_to_images(self, pdf_path: str, output_dir: str, 
                           quality: str = 'medium', format: str = 'PNG') -> List[str]:
        """Convert PDF pages to images with quality options"""
        try:
            loop = asyncio.get_event_loop()
            return await loop.run_in_executor(
                self.executor,
                self._pdf_to_images_sync,
                pdf_path, output_dir, quality, format
            )
        except Exception as e:
            self.logger.error(f"Error converting PDF to images: {e}")
            return []
    
    def _pdf_to_images_sync(self, pdf_path: str, output_dir: str, 
                           quality: str, format: str) -> List[str]:
        """Synchronous PDF to images conversion"""
        try:
            # Verify PDF exists
            if not os.path.exists(pdf_path):
                self.logger.error(f"PDF file not found: {pdf_path}")
                return []
                
            doc = fitz.open(pdf_path)
            image_paths = []
            dpi = self.config.image_quality.get(quality, 150)
            
            for page_num in range(doc.page_count):
                page = doc[page_num]
                mat = fitz.Matrix(dpi/72, dpi/72)
                pix = page.get_pixmap(matrix=mat)
                
                img_path = os.path.join(output_dir, f"page_{page_num + 1:03d}.{format.lower()}")
                
                if format.upper() == 'PNG':
                    pix.save(img_path)
                else:
                    # Convert to PIL for other formats
                    img_data = pix.tobytes("ppm")
                    img = Image.open(io.BytesIO(img_data))
                    img.save(img_path, format.upper(), quality=95 if format.upper() == 'JPEG' else None)
                
                image_paths.append(img_path)
                pix = None  # Free memory
            
            doc.close()
            return image_paths
            
        except Exception as e:
            self.logger.error(f"Error in _pdf_to_images_sync: {e}")
            return []

    async def excel_to_pdf(self, excel_path: str, output_path: str) -> bool:
        """ENHANCED Excel to PDF conversion with proper formatting and layout"""
        try:
            # Verify file exists
            if not os.path.exists(excel_path):
                self.logger.error(f"Excel file not found: {excel_path}")
                return False
                
            # Method 1: Try LibreOffice first (best for preserving formatting)
            if await self._convert_excel_with_libreoffice(excel_path, output_path):
                return True
            
            # Method 2: Enhanced conversion with proper table formatting
            return await self._convert_excel_enhanced(excel_path, output_path)
            
        except Exception as e:
            self.logger.error(f"Error converting Excel to PDF: {e}")
            return False

    async def _convert_excel_with_libreoffice(self, excel_path: str, output_path: str) -> bool:
        """Convert Excel using LibreOffice - preserves formatting best"""
        try:
            # Try different LibreOffice commands
            commands = [
                ['libreoffice', '--headless', '--convert-to', 'pdf', '--outdir', os.path.dirname(output_path), excel_path],
                ['soffice', '--headless', '--convert-to', 'pdf', '--outdir', os.path.dirname(output_path), excel_path]
            ]
            
            for cmd in commands:
                try:
                    result = await asyncio.create_subprocess_exec(
                        *cmd, 
                        stdout=asyncio.subprocess.PIPE, 
                        stderr=asyncio.subprocess.PIPE
                    )
                    stdout, stderr = await result.communicate()
                    
                    # LibreOffice creates file with same name but .pdf extension
                    expected_output = os.path.join(
                        os.path.dirname(output_path),
                        os.path.splitext(os.path.basename(excel_path))[0] + '.pdf'
                    )
                    
                    if os.path.exists(expected_output):
                        if expected_output != output_path:
                            shutil.move(expected_output, output_path)
                        return True
                except FileNotFoundError:
                    continue
            
            return False
            
        except Exception as e:
            self.logger.warning(f"LibreOffice Excel conversion failed: {e}")
            return False

    async def _convert_excel_enhanced(self, excel_path: str, output_path: str) -> bool:
        """ENHANCED Excel to PDF with proper table formatting"""
        try:
            loop = asyncio.get_event_loop()
            return await loop.run_in_executor(
                self.executor,
                self._excel_enhanced_sync,
                excel_path, output_path
            )
        except Exception as e:
            self.logger.error(f"Enhanced Excel conversion failed: {e}")
            return False

    def _excel_enhanced_sync(self, excel_path: str, output_path: str) -> bool:
        """ENHANCED synchronous Excel to PDF conversion with proper formatting"""
        try:
            # Verify file exists
            if not os.path.exists(excel_path):
                self.logger.error(f"Excel file not found: {excel_path}")
                return False
                
            # Read Excel file with openpyxl to preserve formatting
            wb = openpyxl.load_workbook(excel_path, data_only=False)
            
            # Create PDF
            doc = SimpleDocTemplate(output_path, pagesize=A4, 
                                  rightMargin=30, leftMargin=30, 
                                  topMargin=30, bottomMargin=30)
            elements = []
            styles = getSampleStyleSheet()
            
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                
                # Add sheet title
                title = Paragraph(f"<b>Sheet: {sheet_name}</b>", styles['Heading1'])
                elements.append(title)
                elements.append(Spacer(1, 12))
                
                # Get actual data range (not just used cells)
                max_row = ws.max_row
                max_col = ws.max_column
                
                if max_row > 0 and max_col > 0:
                    # Create table data with proper formatting
                    table_data = []
                    
                    # Limit to reasonable size for PDF
                    display_rows = min(max_row, 50)
                    display_cols = min(max_col, 10)
                    
                    for row_idx in range(1, display_rows + 1):
                        row_data = []
                        for col_idx in range(1, display_cols + 1):
                            cell = ws.cell(row=row_idx, column=col_idx)
                            
                            # Get cell value with proper formatting
                            if cell.value is not None:
                                # Handle different data types
                                if isinstance(cell.value, (int, float)):
                                    if cell.number_format and cell.number_format != 'General':
                                        # Try to format numbers according to Excel format
                                        cell_text = str(cell.value)
                                    else:
                                        cell_text = str(cell.value)
                                elif isinstance(cell.value, datetime):
                                    cell_text = cell.value.strftime('%Y-%m-%d %H:%M:%S')
                                else:
                                    cell_text = str(cell.value)
                            else:
                                cell_text = ""
                            
                            # Limit cell content length for PDF
                            if len(cell_text) > 30:
                                cell_text = cell_text[:27] + "..."
                            
                            row_data.append(cell_text)
                        
                        table_data.append(row_data)
                    
                    if table_data:
                        # Create table with proper styling
                        table = Table(table_data, hAlign='LEFT')
                        
                        # Apply table styling
                        table_style = TableStyle([
                            # Basic table structure
                            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),  # Header row
                            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                            ('FONTSIZE', (0, 0), (-1, 0), 10),
                            ('FONTSIZE', (0, 1), (-1, -1), 8),
                            ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
                            
                            # Grid lines
                            ('GRID', (0, 0), (-1, -1), 1, colors.black),
                            ('LINEBELOW', (0, 0), (-1, 0), 2, colors.black),
                            
                            # Alternating row colors
                            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
                            
                            # Cell padding
                            ('LEFTPADDING', (0, 0), (-1, -1), 3),
                            ('RIGHTPADDING', (0, 0), (-1, -1), 3),
                            ('TOPPADDING', (0, 0), (-1, -1), 3),
                            ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
                        ])
                        
                        table.setStyle(table_style)
                        elements.append(table)
                        
                        # Add info if data was truncated
                        if max_row > 50 or max_col > 10:
                            info_text = f"(Showing {display_rows} of {max_row} rows, {display_cols} of {max_col} columns)"
                            elements.append(Spacer(1, 6))
                            elements.append(Paragraph(f"<i>{info_text}</i>", styles['Normal']))
                    else:
                        elements.append(Paragraph("<i>(Empty sheet)</i>", styles['Normal']))
                else:
                    elements.append(Paragraph("<i>(Empty sheet)</i>", styles['Normal']))
                
                # Add space between sheets
                elements.append(Spacer(1, 24))
            
            if elements:
                doc.build(elements)
                return True
            return False
            
        except Exception as e:
            self.logger.error(f"Error in _excel_enhanced_sync: {e}")
            return False

    async def word_to_pdf(self, docx_path: str, output_path: str) -> bool:
        """Convert Word document to PDF with multiple methods"""
        try:
            # Verify file exists
            if not os.path.exists(docx_path):
                self.logger.error(f"Word file not found: {docx_path}")
                return False
                
            # Method 1: Try LibreOffice (most reliable)
            if await self._convert_with_libreoffice(docx_path, output_path):
                return True
            
            # Method 2: Try pandoc
            if await self._convert_with_pandoc(docx_path, output_path):
                return True
                
            # Method 3: Basic conversion using python-docx + reportlab
            return await self._convert_docx_basic(docx_path, output_path)
            
        except Exception as e:
            self.logger.error(f"Error converting Word to PDF: {e}")
            return False
    
    async def _convert_with_libreoffice(self, input_path: str, output_path: str) -> bool:
        """Convert using LibreOffice"""
        try:
            # Try different LibreOffice commands
            commands = [
                ['libreoffice', '--headless', '--convert-to', 'pdf', '--outdir', os.path.dirname(output_path), input_path],
                ['soffice', '--headless', '--convert-to', 'pdf', '--outdir', os.path.dirname(output_path), input_path]
            ]
            
            for cmd in commands:
                try:
                    result = await asyncio.create_subprocess_exec(
                        *cmd, 
                        stdout=asyncio.subprocess.PIPE, 
                        stderr=asyncio.subprocess.PIPE
                    )
                    stdout, stderr = await result.communicate()
                    
                    # LibreOffice creates file with same name but .pdf extension
                    expected_output = os.path.join(
                        os.path.dirname(output_path),
                        os.path.splitext(os.path.basename(input_path))[0] + '.pdf'
                    )
                    
                    if os.path.exists(expected_output):
                        if expected_output != output_path:
                            shutil.move(expected_output, output_path)
                        return True
                except FileNotFoundError:
                    continue
            
            return False
            
        except Exception as e:
            self.logger.warning(f"LibreOffice conversion failed: {e}")
            return False
    
    async def _convert_with_pandoc(self, input_path: str, output_path: str) -> bool:
        """Convert using Pandoc"""
        try:
            cmd = ['pandoc', input_path, '-o', output_path]
            result = await asyncio.create_subprocess_exec(
                *cmd, stdout=asyncio.subprocess.PIPE, stderr=asyncio.subprocess.PIPE
            )
            stdout, stderr = await result.communicate()
            return os.path.exists(output_path)
        except Exception as e:
            self.logger.warning(f"Pandoc conversion failed: {e}")
            return False
    
    async def _convert_docx_basic(self, docx_path: str, output_path: str) -> bool:
        """Basic DOCX to PDF conversion"""
        try:
            loop = asyncio.get_event_loop()
            return await loop.run_in_executor(
                self.executor,
                self._docx_basic_sync,
                docx_path, output_path
            )
        except Exception as e:
            self.logger.error(f"Basic DOCX conversion failed: {e}")
            return False
    
    def _docx_basic_sync(self, docx_path: str, output_path: str) -> bool:
        """Synchronous basic DOCX conversion"""
        try:
            # Verify file exists
            if not os.path.exists(docx_path):
                self.logger.error(f"DOCX file not found: {docx_path}")
                return False
                
            doc = DocxDocument(docx_path)
            
            # Create PDF
            pdf_doc = SimpleDocTemplate(output_path, pagesize=A4)
            elements = []
            styles = getSampleStyleSheet()
            
            for paragraph in doc.paragraphs:
                if paragraph.text.strip():
                    # Clean text and handle special characters
                    text = paragraph.text.replace('\n', '<br/>')
                    text = text.replace('\t', '    ')
                    p = Paragraph(text, styles['Normal'])
                    elements.append(p)
                    elements.append(Spacer(1, 12))
            
            if elements:
                pdf_doc.build(elements)
                return True
            return False
            
        except Exception as e:
            self.logger.error(f"Error in _docx_basic_sync: {e}")
            return False
    
    def enhance_image(self, image_path: str, enhancement_type: str, strength: float = 1.2) -> str:
        """Enhanced image processing with more options - FIXED filename handling"""
        try:
            if not os.path.exists(image_path):
                self.logger.error(f"Image file not found: {image_path}")
                return image_path
                
            img = Image.open(image_path)
            original_format = img.format or 'JPEG'
            
            if enhancement_type == 'brightness':
                enhancer = ImageEnhance.Brightness(img)
                img = enhancer.enhance(strength)
            elif enhancement_type == 'contrast':
                enhancer = ImageEnhance.Contrast(img)
                img = enhancer.enhance(strength)
            elif enhancement_type == 'sharpness':
                enhancer = ImageEnhance.Sharpness(img)
                img = enhancer.enhance(strength)
            elif enhancement_type == 'color':
                enhancer = ImageEnhance.Color(img)
                img = enhancer.enhance(strength)
            elif enhancement_type == 'auto_enhance':
                # Auto enhance: apply multiple enhancements
                img = ImageEnhance.Brightness(img).enhance(1.1)
                img = ImageEnhance.Contrast(img).enhance(1.2)
                img = ImageEnhance.Sharpness(img).enhance(1.1)
            elif enhancement_type == 'grayscale':
                img = img.convert('L').convert('RGB')
            elif enhancement_type == 'blur':
                img = img.filter(ImageFilter.GaussianBlur(radius=2))
            
            # Save enhanced image with cleaner filename
            base_dir = os.path.dirname(image_path)
            base_name = os.path.splitext(os.path.basename(image_path))[0]
            # Clean base name if it's too long
            if len(base_name) > 30:
                base_name = f"enhanced_{uuid.uuid4().hex[:8]}"
            else:
                base_name = f"{base_name}_enh"
            
            enhanced_path = os.path.join(base_dir, f"{base_name}_{enhancement_type}.jpg")
            img.save(enhanced_path, format='JPEG', quality=95)
            return enhanced_path
            
        except Exception as e:
            self.logger.error(f"Error enhancing image: {e}")
            return image_path
